#!/usr/bin/env python3
"""Normalize a bucket earnings event workbook into a backend-friendly CSV table."""

from __future__ import annotations

import argparse
import csv
from datetime import date, datetime
from pathlib import Path
from typing import Iterable
import re

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
Buckets_DIR = ROOT / "buckets"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Normalize a wide bucket events workbook into a long-form CSV table."
    )
    parser.add_argument("workbook", type=Path, help="Path to the source .xlsx workbook.")
    parser.add_argument(
        "--sheet",
        default=None,
        help="Optional sheet name. Defaults to the first worksheet.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Optional output CSV path. Defaults next to the workbook as *_event_returns.csv.",
    )
    parser.add_argument(
        "--as-of-date",
        default=None,
        help="Optional YYYY-MM-DD stamp for when the workbook was processed.",
    )
    return parser.parse_args()


def normalize_instrument(value: object) -> str:
    return str(value or "").strip()


def normalize_bucket_symbol(value: object) -> str:
    return str(value or "").strip()


def bucket_fs_from_symbol(bucket_symbol: str, workbook: Path) -> str:
    if bucket_symbol:
        return bucket_symbol.lstrip(".")
    return workbook.stem.removesuffix("_events")


def parse_excel_date(value: object) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def numeric_or_none(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def parse_iso_date(value: str) -> date | None:
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def instrument_to_ticker(instrument: str) -> str:
    text = instrument.strip()
    if not text:
        return ""
    return text.split()[0]


def parse_note_sections(text: str) -> list[tuple[str, str]]:
    parts = re.split(r"(^##\s+.+$)", text, flags=re.M)
    sections: list[tuple[str, str]] = []
    for idx in range(1, len(parts), 2):
        header = parts[idx].strip()[3:].strip()
        body = parts[idx + 1]
        sections.append((header, body))
    return sections


def parse_field(body: str, field_name: str) -> str | None:
    match = re.search(rf"^{re.escape(field_name)}:\s*(.+?)\s*$", body, flags=re.M)
    if not match:
        return None
    value = match.group(1).strip()
    return value or None


def load_company_note_matches(bucket_fs: str, event_ticker: str) -> list[dict[str, str]]:
    ticker = instrument_to_ticker(event_ticker)
    if not ticker:
        return []
    note_path = Buckets_DIR / bucket_fs / ticker / f"{ticker}.md"
    if not note_path.exists():
        return []
    text = note_path.read_text(encoding="utf-8", errors="ignore")
    matches: list[dict[str, str]] = []
    for header, body in parse_note_sections(text):
        calendar_quarter = parse_field(body, "calendar_quarter") or header
        event_date = parse_field(body, "event_date") or ""
        if not calendar_quarter:
            continue
        matches.append(
            {
                "calendar_quarter": calendar_quarter,
                "note_heading": header,
                "note_event_date": event_date,
                "note_path": str(note_path.relative_to(ROOT)),
            }
        )
    return matches


def match_note_block(bucket_fs: str, event_ticker: str, report_date: str, tolerance_days: int = 2) -> dict[str, str]:
    report_day = parse_iso_date(report_date)
    best_match: dict[str, str] | None = None
    best_distance: int | None = None
    for candidate in load_company_note_matches(bucket_fs, event_ticker):
        note_day = parse_iso_date(candidate["note_event_date"])
        if note_day is None or report_day is None:
            continue
        distance = abs((note_day - report_day).days)
        if best_distance is None or distance < best_distance:
            best_distance = distance
            best_match = candidate
    if best_match is None:
        return {
            "calendar_quarter": "",
            "matched_note_event_date": "",
            "matched_note_heading": "",
            "matched_note_path": "",
            "match_quality": "unmatched",
        }
    if best_distance is None or best_distance > tolerance_days:
        return {
            "calendar_quarter": "",
            "matched_note_event_date": "",
            "matched_note_heading": "",
            "matched_note_path": "",
            "match_quality": "unmatched",
        }
    match_quality = "exact_date" if best_distance == 0 else "near_date"
    return {
        "calendar_quarter": best_match["calendar_quarter"],
        "matched_note_event_date": best_match["note_event_date"],
        "matched_note_heading": best_match["note_heading"],
        "matched_note_path": best_match["note_path"],
        "match_quality": match_quality,
    }


def iter_rows(ws) -> Iterable[list[object]]:
    for row in ws.iter_rows(values_only=True):
        yield list(row)


def normalize_workbook(workbook_path: Path, sheet_name: str | None, as_of_date: str | None) -> list[dict[str, object]]:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = list(iter_rows(ws))
    if len(rows) < 2:
        raise ValueError(f"Workbook has insufficient rows: {workbook_path}")

    header = rows[0]
    if len(header) < 4:
        raise ValueError("Expected at least four header columns: bucket, ticker, report_date, instruments...")

    instrument_columns = []
    for column_index, cell_value in enumerate(header[3:], start=4):
        instrument = normalize_instrument(cell_value)
        if instrument:
            instrument_columns.append((column_index, instrument))
    if not instrument_columns:
        raise ValueError("No instrument columns found in workbook header.")

    benchmark_instrument = instrument_columns[0][1]
    output_rows: list[dict[str, object]] = []

    for row_index, row in enumerate(rows[1:], start=2):
        bucket_symbol = normalize_bucket_symbol(row[0] if len(row) >= 1 else None)
        event_ticker = normalize_instrument(row[1] if len(row) >= 2 else None)
        report_date = parse_excel_date(row[2] if len(row) >= 3 else None)

        # Skip metadata rows like the benchmark/index descriptor row.
        if not bucket_symbol or not event_ticker or not report_date:
            continue

        bucket_fs = bucket_fs_from_symbol(bucket_symbol, workbook_path)
        benchmark_return = numeric_or_none(row[instrument_columns[0][0] - 1] if len(row) >= instrument_columns[0][0] else None)
        note_match = match_note_block(bucket_fs, event_ticker, report_date)

        for column_index, instrument in instrument_columns:
            raw_return = numeric_or_none(row[column_index - 1] if len(row) >= column_index else None)
            if raw_return is None:
                continue
            output_rows.append(
                {
                    "as_of_date": as_of_date or date.today().isoformat(),
                    "bucket_symbol": bucket_symbol,
                    "bucket_fs": bucket_fs,
                    "event_ticker": event_ticker,
                    "report_date": report_date,
                    "calendar_quarter": note_match["calendar_quarter"],
                    "benchmark_instrument": benchmark_instrument,
                    "instrument": instrument,
                    "instrument_role": "benchmark" if instrument == benchmark_instrument else "member",
                    "raw_return": raw_return,
                    "relative_return_vs_benchmark": (
                        None if benchmark_return is None else raw_return - benchmark_return
                    ),
                    "matched_note_event_date": note_match["matched_note_event_date"],
                    "matched_note_heading": note_match["matched_note_heading"],
                    "matched_note_path": note_match["matched_note_path"],
                    "match_quality": note_match["match_quality"],
                    "source_workbook": str(workbook_path.relative_to(ROOT)),
                    "source_sheet": ws.title,
                    "source_row": row_index,
                }
            )

    return output_rows


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    fieldnames = [
        "as_of_date",
        "bucket_symbol",
        "bucket_fs",
        "event_ticker",
        "report_date",
        "calendar_quarter",
        "benchmark_instrument",
        "instrument",
        "instrument_role",
        "raw_return",
        "relative_return_vs_benchmark",
        "matched_note_event_date",
        "matched_note_heading",
        "matched_note_path",
        "match_quality",
        "source_workbook",
        "source_sheet",
        "source_row",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook
    if not workbook_path.is_absolute():
        workbook_path = (ROOT / workbook_path).resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    output_path = args.output
    if output_path is None:
        stem = workbook_path.stem
        if stem.endswith("_events"):
            stem = stem[: -len("_events")]
        output_path = workbook_path.with_name(f"{stem}_event_returns.csv")
    elif not output_path.is_absolute():
        output_path = (ROOT / output_path).resolve()

    rows = normalize_workbook(workbook_path, args.sheet, args.as_of_date)
    write_csv(output_path, rows)
    print(f"Wrote {len(rows)} normalized rows to {output_path}")


if __name__ == "__main__":
    main()

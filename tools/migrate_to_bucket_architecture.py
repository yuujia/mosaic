"""Migrate Mosaic into bucket architecture with dry-run/apply safety modes."""

from __future__ import annotations

import argparse
import subprocess
import shutil
import sys
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable


@dataclass(frozen=True)
class CompanyRow:
    bucket: str
    ticker: str


@dataclass
class Stats:
    buckets_created: int = 0
    companies_created: int = 0
    notes_copied: int = 0
    notes_missing: int = 0
    skipped_conflicts: int = 0


def is_subpath(path: Path, parent: Path) -> bool:
    try:
        path.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def maybe_relaunch_from_temp(args: argparse.Namespace, root: Path) -> int | None:
    if not args.apply:
        return None
    script_path = Path(__file__).resolve()
    if not is_subpath(script_path, root):
        return None

    temp_runner_dir = Path(tempfile.gettempdir()) / "mosaic_migration_runner"
    temp_runner_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    temp_script = temp_runner_dir / f"migrate_to_bucket_architecture_{ts}.py"
    shutil.copy2(script_path, temp_script)

    cmd = [sys.executable, str(temp_script), "--apply", "--root", str(root)]
    if args.force:
        cmd.append("--force")

    print(f"INFO: Running apply from temporary runner to avoid file locks: {temp_script}")
    completed = subprocess.run(cmd)
    return int(completed.returncode)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Move current Mosaic folder to timestamped legacy, rebuild a fresh bucket-based skeleton, "
            "and migrate portfolio + company notes."
        )
    )
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument(
        "--dry-run",
        action="store_true",
        help="Print planned operations only (default).",
    )
    mode.add_argument(
        "--apply",
        action="store_true",
        help="Execute the migration operations.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Allow overwriting destination files if they already exist.",
    )
    parser.add_argument(
        "--root",
        type=Path,
        default=Path(__file__).resolve().parents[1],
        help="Path to current Mosaic root folder (defaults to script parent).",
    )
    return parser.parse_args()


def normalize_ticker(raw: str) -> str:
    value = " ".join((raw or "").strip().upper().split())
    if not value:
        return ""
    token = value.split(" ", 1)[0]
    for ch in '<>:"/\\|?*':
        token = token.replace(ch, "_")
    return token.strip(" .")


def normalize_bucket(raw: str) -> str:
    return (raw or "").strip()


def is_ignored_bucket(bucket: str) -> bool:
    normalized = bucket.strip().upper()
    return normalized in {"", "N/A", "NA", "#N/A", "#N/A N/A", "NONE", "NULL"}


def find_header_index(headers: list[str], candidates: Iterable[str]) -> int:
    lookup = {str(name).strip().lower(): idx for idx, name in enumerate(headers)}
    for candidate in candidates:
        idx = lookup.get(candidate.strip().lower())
        if idx is not None:
            return idx
    raise ValueError(f"Missing required column. Tried: {', '.join(candidates)}")


def load_portfolio(path: Path) -> tuple[list[str], list[CompanyRow], set[str]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - runtime dependency guard
        raise SystemExit(
            "Missing dependency: openpyxl. Install requirements first (e.g., `pip install -r requirements.txt`)."
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        wb.close()
        raise ValueError(f"Portfolio workbook is empty: {path}")
    headers = [str(x or "").strip() for x in header_row]
    bucket_idx = find_header_index(headers, ["Bucket Symbol", "Bucket", "bucket_symbol"])
    ticker_idx = find_header_index(headers, ["Ticker", "ticker", "Ticker Symbol"])

    companies: list[CompanyRow] = []
    unique_buckets: set[str] = set()
    seen_company: set[tuple[str, str]] = set()

    for row in rows:
        bucket = normalize_bucket(str(row[bucket_idx] or ""))
        ticker = normalize_ticker(str(row[ticker_idx] or ""))
        if is_ignored_bucket(bucket) or not ticker:
            continue
        unique_buckets.add(bucket)
        key = (bucket, ticker)
        if key in seen_company:
            continue
        seen_company.add(key)
        companies.append(CompanyRow(bucket=bucket, ticker=ticker))

    wb.close()
    return headers, companies, unique_buckets


def print_plan(label: str, src: Path | None = None, dst: Path | None = None) -> None:
    if src is not None and dst is not None:
        print(f"PLAN {label}: {src} -> {dst}")
    elif dst is not None:
        print(f"PLAN {label}: {dst}")
    else:
        print(f"PLAN {label}")


def ensure_dir(path: Path, force: bool, stats: Stats) -> bool:
    if path.exists():
        if not path.is_dir():
            if not force:
                print(f"SKIP CONFLICT (exists file, needs dir): {path}")
                stats.skipped_conflicts += 1
                return False
            path.unlink()
            path.mkdir(parents=True, exist_ok=True)
            print(f"APPLY MKDIR (force replace file): {path}")
            return True
        print(f"SKIP EXISTS (dir): {path}")
        stats.skipped_conflicts += 1
        return False
    path.mkdir(parents=True, exist_ok=True)
    print(f"APPLY MKDIR: {path}")
    return True


def copy_file(src: Path, dst: Path, force: bool, stats: Stats) -> bool:
    if dst.exists() and not force:
        print(f"SKIP CONFLICT (file exists): {dst}")
        stats.skipped_conflicts += 1
        return False
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    print(f"APPLY COPY: {src} -> {dst}")
    return True


def resolve_bucket_kpi_template(root: Path) -> Path:
    candidates = [
        root / "tools" / "templates" / "BUCKET_kpi_template.xlsx",
        root / "tools" / "templates" / "BUCKET_kpi_TEMPLATE.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    expected = " or ".join(str(p) for p in candidates)
    raise FileNotFoundError(f"Missing KPI template file: {expected}")


def main() -> int:
    args = parse_args()
    apply_mode = bool(args.apply)
    dry_run = not apply_mode

    root = args.root.resolve()
    relaunch_code = maybe_relaunch_from_temp(args, root)
    if relaunch_code is not None:
        return relaunch_code

    parent = root.parent
    if root.name != "Mosaic":
        print(f"WARNING: root folder name is '{root.name}', expected 'Mosaic'. Continuing anyway.")

    if not root.exists():
        print(f"ERROR: Root does not exist: {root}")
        return 2

    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    date_part = datetime.now().strftime("%Y-%m-%d")
    legacy_root = parent / f"{root.name}__LEGACY_{date_part}_{ts}"
    new_root = parent / "Mosaic"

    source_portfolio = root / "00_config" / "active_portfolio.xlsx"
    if not source_portfolio.exists():
        print(f"ERROR: Missing portfolio file: {source_portfolio}")
        return 2
    try:
        source_kpi_template = resolve_bucket_kpi_template(root)
    except FileNotFoundError as exc:
        print(f"ERROR: {exc}")
        return 2

    try:
        _headers, companies, unique_buckets = load_portfolio(source_portfolio)
    except Exception as exc:
        print(f"ERROR: Failed to read portfolio workbook: {exc}")
        return 2

    companies_by_bucket: dict[str, list[str]] = {}
    for row in companies:
        companies_by_bucket.setdefault(row.bucket, []).append(row.ticker)

    print(f"Mode: {'APPLY' if apply_mode else 'DRY RUN'}")
    print(f"Current root: {root}")
    print(f"Planned legacy root: {legacy_root}")
    print(f"Planned fresh root: {new_root}")
    print("")
    print_plan("[MOVE ROOT]", src=root, dst=legacy_root)
    print_plan("[MKDIR]", dst=new_root)
    print_plan("[MKDIR]", dst=new_root / "00_config")
    print_plan("[MKDIR]", dst=new_root / "buckets")
    print_plan("[MKDIR]", dst=new_root / "tools")
    print_plan(
        "[COPY FILE]",
        src=source_kpi_template,
        dst=new_root / "tools" / "templates" / source_kpi_template.name,
    )
    print_plan(
        "[COPY FILE]",
        src=legacy_root / "00_config" / "active_portfolio.xlsx",
        dst=new_root / "00_config" / "active_portfolio.xlsx",
    )

    for bucket in sorted(unique_buckets):
        bucket_dir = new_root / "buckets" / bucket
        kpi_path = bucket_dir / f"{bucket}_kpi.xlsx"
        print_plan("[MKDIR]", dst=bucket_dir)
        print_plan("[COPY FILE]", src=source_kpi_template, dst=kpi_path)

    for bucket in sorted(companies_by_bucket):
        for ticker in sorted(companies_by_bucket[bucket]):
            ticker_root = new_root / "buckets" / bucket / ticker
            source_dir = ticker_root / "source"
            print_plan("[MKDIR]", dst=ticker_root)
            print_plan("[MKDIR]", dst=source_dir)
            note_src_planned = legacy_root / "01_Companies" / bucket / ticker / f"{ticker}.docx"
            note_src_now = root / "01_Companies" / bucket / ticker / f"{ticker}.docx"
            note_dst = ticker_root / f"{ticker}.docx"
            if note_src_now.exists():
                print_plan("[COPY FILE]", src=note_src_planned, dst=note_dst)
            else:
                print(f"PLAN [MISSING NOTE]: {note_src_planned}")

    if dry_run:
        notes_found = sum(
            1
            for row in companies
            if (root / "01_Companies" / row.bucket / row.ticker / f"{row.ticker}.docx").exists()
        )
        notes_missing = len(companies) - notes_found
        print("")
        print("Summary (planned):")
        print(f"- legacy folder path created: {legacy_root}")
        print(f"- number of buckets created: {len(unique_buckets)}")
        print(f"- number of companies created: {len(companies)}")
        print(f"- notes found+copied vs missing: {notes_found} copied, {notes_missing} missing")
        print("- files/folders skipped due to conflicts: 0 (projected; real count shown in --apply)")
        return 0

    stats = Stats()

    if legacy_root.exists():
        if not args.force:
            print(f"SKIP CONFLICT (legacy target exists): {legacy_root}")
            stats.skipped_conflicts += 1
            print("")
            print("Summary:")
            print(f"- legacy folder path created: {legacy_root} (not created)")
            print("- number of buckets created: 0")
            print("- number of companies created: 0")
            print("- notes found+copied vs missing: 0 copied, 0 missing")
            print(f"- files/folders skipped due to conflicts: {stats.skipped_conflicts}")
            return 1
        print(f"ERROR: --force cannot safely merge into existing legacy target: {legacy_root}")
        return 2

    try:
        root.rename(legacy_root)
    except PermissionError as exc:
        print(f"ERROR: Unable to rename root folder due to file lock: {exc}")
        print("Close Explorer windows, terminals, editors, and Dropbox handles using this folder, then retry.")
        return 2
    except OSError as exc:
        print(f"ERROR: Unable to rename root folder: {exc}")
        return 2
    print(f"APPLY MOVE ROOT: {root} -> {legacy_root}")

    created_root = ensure_dir(new_root, args.force, stats)
    created_cfg = ensure_dir(new_root / "00_config", args.force, stats)
    created_buckets = ensure_dir(new_root / "buckets", args.force, stats)
    ensure_dir(new_root / "tools", args.force, stats)
    ensure_dir(new_root / "tools" / "templates", args.force, stats)
    copy_file(
        source_kpi_template,
        new_root / "tools" / "templates" / source_kpi_template.name,
        args.force,
        stats,
    )
    if created_root or created_cfg or created_buckets:
        pass

    legacy_portfolio = legacy_root / "00_config" / "active_portfolio.xlsx"
    if legacy_portfolio.exists():
        copy_file(legacy_portfolio, new_root / "00_config" / "active_portfolio.xlsx", args.force, stats)
    else:
        print(f"WARNING: Missing legacy portfolio after move: {legacy_portfolio}")

    for bucket in sorted(unique_buckets):
        bucket_dir = new_root / "buckets" / bucket
        kpi_path = bucket_dir / f"{bucket}_kpi.xlsx"

        bucket_created = ensure_dir(bucket_dir, args.force, stats)
        if bucket_created:
            stats.buckets_created += 1

        copy_file(source_kpi_template, kpi_path, args.force, stats)

    for row in companies:
        ticker_root = new_root / "buckets" / row.bucket / row.ticker
        source_dir = ticker_root / "source"

        company_created = ensure_dir(ticker_root, args.force, stats)
        ensure_dir(source_dir, args.force, stats)
        if company_created:
            stats.companies_created += 1

        legacy_note = legacy_root / "01_Companies" / row.bucket / row.ticker / f"{row.ticker}.docx"
        note_dst = ticker_root / f"{row.ticker}.docx"
        if legacy_note.exists():
            copied = copy_file(legacy_note, note_dst, args.force, stats)
            if copied:
                stats.notes_copied += 1
        else:
            stats.notes_missing += 1
            print(f"MISSING NOTE: {legacy_note}")

    print("")
    print("Summary:")
    print(f"- legacy folder path created: {legacy_root}")
    print(f"- number of buckets created: {stats.buckets_created}")
    print(f"- number of companies created: {stats.companies_created}")
    print(f"- notes found+copied vs missing: {stats.notes_copied} copied, {stats.notes_missing} missing")
    print(f"- files/folders skipped due to conflicts: {stats.skipped_conflicts}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

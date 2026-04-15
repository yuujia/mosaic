"""Export bucket KPI workbooks into text-friendly CSV mirrors for GitHub-based agents."""

from __future__ import annotations

import argparse
import csv
import time
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
BUCKETS_DIR = ROOT / "buckets"
EXPORT_DIR = ROOT / "00_config" / "kpi_exports"
REQUIRED_HEADERS = ("company", "calendar_quarter", "type", "kpi", "value")
OPTIONAL_HEADERS = ("yoy", "yoy_vs_last_q", "yoy_vs_last_y", "qoq")
HEADER_ALIASES = {
    "company": "company",
    "calendar_quarter": "calendar_quarter",
    "type": "type",
    "kpi": "kpi",
    "value": "value",
    "yoy": "yoy",
    "yoy_vs_last_q": "yoy_vs_last_q",
    "yoy_vs_last_y": "yoy_vs_last_y",
    "qoq": "qoq",
    "status": "period_status",
    "period_status": "period_status",
    "period status": "period_status",
}


def load_openpyxl_workbook(workbook_path: Path):
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - runtime dependency guard
        raise SystemExit(
            "Missing dependency: openpyxl. Install requirements first (e.g., `pip install -r requirements.txt`)."
        ) from exc

    return load_workbook(workbook_path, data_only=True)


def find_normalized_header_columns(ws) -> dict[str, int] | None:
    header_map: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        value = cell.value
        if value is None:
            continue
        normalized = str(value).strip().lower()
        canonical_name = HEADER_ALIASES.get(normalized)
        if canonical_name is not None:
            header_map[canonical_name] = col_idx
    if not all(header in header_map for header in REQUIRED_HEADERS):
        return None
    return header_map


def export_workbook(workbook_path: Path, export_dir: Path, bucket_fs: str | None = None) -> Path | None:
    wb = load_openpyxl_workbook(workbook_path)
    rows = []

    for ws in wb.worksheets:
        header_map = find_normalized_header_columns(ws)
        if header_map is None:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            company = row[header_map["company"] - 1]
            quarter = row[header_map["calendar_quarter"] - 1]
            record_type = row[header_map["type"] - 1]
            kpi = row[header_map["kpi"] - 1]
            value = row[header_map["value"] - 1]
            period_status = row[header_map["period_status"] - 1] if "period_status" in header_map else None
            yoy = row[header_map["yoy"] - 1] if "yoy" in header_map else None
            yoy_vs_last_q = row[header_map["yoy_vs_last_q"] - 1] if "yoy_vs_last_q" in header_map else None
            yoy_vs_last_y = row[header_map["yoy_vs_last_y"] - 1] if "yoy_vs_last_y" in header_map else None
            qoq = row[header_map["qoq"] - 1] if "qoq" in header_map else None
            if not (company and quarter and kpi and value not in (None, "")):
                continue

            quarter_text = quarter.date().isoformat() if hasattr(quarter, "date") else str(quarter)
            rows.append(
                {
                    "sheet": ws.title,
                    "company": str(company),
                    "calendar_quarter": quarter_text,
                    "type": "" if record_type is None else str(record_type),
                    "period_status": "" if period_status is None else str(period_status),
                    "kpi": str(kpi),
                    "value": value,
                    "yoy": yoy,
                    "yoy_vs_last_q": yoy_vs_last_q,
                    "yoy_vs_last_y": yoy_vs_last_y,
                    "qoq": qoq,
                }
            )

    if not rows:
        return None

    export_dir.mkdir(parents=True, exist_ok=True)
    output_stem = f"{bucket_fs}_kpi" if bucket_fs else workbook_path.stem
    output_path = export_dir / f"{output_stem}.csv"
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "sheet",
                "company",
                "calendar_quarter",
                "type",
                "period_status",
                "kpi",
                "value",
                "yoy",
                "yoy_vs_last_q",
                "yoy_vs_last_y",
                "qoq",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)
    return output_path


def is_bucket_kpi_workbook(path: Path, bucket_fs: str) -> bool:
    return (
        path.is_file()
        and path.suffix.lower() == ".xlsx"
        and not path.name.startswith(".")
        and path.stem.lower() == f"{bucket_fs.lower()}_kpi"
    )


def find_bucket_workbook(bucket_dir: Path) -> Path | None:
    if not bucket_dir.exists():
        return None
    matches = [
        path
        for path in bucket_dir.iterdir()
        if is_bucket_kpi_workbook(path, bucket_dir.name)
    ]
    if not matches:
        return None
    return max(matches, key=lambda path: path.stat().st_mtime)


def iter_bucket_workbooks(bucket: str | None) -> list[tuple[str, Path]]:
    if bucket:
        bucket_fs = bucket.lstrip(".")
        workbook = find_bucket_workbook(BUCKETS_DIR / bucket_fs)
        return [(bucket_fs, workbook)] if workbook else []

    workbooks: list[tuple[str, Path]] = []
    for bucket_dir in sorted(path for path in BUCKETS_DIR.iterdir() if path.is_dir()):
        workbook = find_bucket_workbook(bucket_dir)
        if workbook:
            workbooks.append((bucket_dir.name, workbook))
    return workbooks


def export_workbooks(workbooks: list[tuple[str, Path]], export_dir: Path) -> int:
    written = 0
    for bucket_fs, workbook_path in workbooks:
        output_path = export_workbook(workbook_path, export_dir, bucket_fs=bucket_fs)
        if output_path is not None:
            written += 1
            print(output_path.relative_to(ROOT), flush=True)
    return written


def snapshot_mtimes(workbooks: list[tuple[str, Path]]) -> dict[Path, float]:
    return {workbook_path: workbook_path.stat().st_mtime for _, workbook_path in workbooks}


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Export Mosaic bucket KPI workbooks into text-friendly CSV mirrors."
    )
    parser.add_argument(
        "--bucket",
        help="Bucket ID or dotted bucket symbol to export. Defaults to all bucket KPI workbooks.",
    )
    parser.add_argument(
        "--output-dir",
        default=str(EXPORT_DIR),
        help="Directory for CSV exports. Default: 00_config/kpi_exports",
    )
    parser.add_argument(
        "--watch",
        action="store_true",
        help="Keep running and refresh CSV exports when matching KPI workbooks are saved.",
    )
    parser.add_argument(
        "--interval",
        type=float,
        default=2.0,
        help="Polling interval in seconds for --watch. Default: 2.0",
    )
    args = parser.parse_args()

    export_dir = Path(args.output_dir)
    workbooks = iter_bucket_workbooks(args.bucket)
    if not workbooks:
        raise SystemExit("No matching bucket KPI workbooks found.")

    written = export_workbooks(workbooks, export_dir)
    if written == 0:
        raise SystemExit("No normalized KPI rows found to export.")

    if args.watch:
        mtimes = snapshot_mtimes(workbooks)
        print("Watching KPI workbooks for changes. Press Ctrl+C to stop.", flush=True)
        while True:
            time.sleep(args.interval)
            current_workbooks = iter_bucket_workbooks(args.bucket)
            current_mtimes = snapshot_mtimes(current_workbooks)
            changed_paths = [
                path
                for path, mtime in current_mtimes.items()
                if mtimes.get(path) != mtime
            ]
            if changed_paths:
                changed_workbooks = [
                    (bucket_fs, workbook_path)
                    for bucket_fs, workbook_path in current_workbooks
                    if workbook_path in changed_paths
                ]
                export_workbooks(changed_workbooks, export_dir)
                mtimes = current_mtimes
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
